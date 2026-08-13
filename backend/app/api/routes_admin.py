"""
routes_admin.py — Endpoints exclusivos del rol Admin.

Cubre:
  POST /admin/importar-declarantes   Carga masiva desde Excel (.xlsx)
  GET  /admin/plantilla-declarantes  Descarga la plantilla Excel estándar
"""
import io

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.permisos import requiere_rol
from app.db.session import get_db
from app.models.usuario import RolUsuario, Usuario
from app.services import declarante_service as svc

router = APIRouter(
    prefix="/admin",
    tags=["admin"],
    dependencies=[Depends(requiere_rol(RolUsuario.ADMIN))],
)

# Columnas de la plantilla — única fuente de verdad (eliminada la duplicación
# con declarante_service.COLUMNAS_OBLIGATORIAS; la plantilla puede tener más
# columnas que las mínimas obligatorias).
COLUMNAS_PLANTILLA = [
    "nit",
    "digito_verificacion",
    "primer_apellido",
    "primer_nombre",
    "actividad_economica",      # empleado | independiente | rentista | otro
    "patrimonio_bruto_2025",    # opcional — número sin puntos
    "pasivos_2025",             # opcional
    "patrimonio_bruto_2024",    # opcional — para renta presuntiva
    "pasivos_2024",             # opcional
]


# ── Schemas de respuesta ───────────────────────────────────────────────────────

class FilaResultado(BaseModel):
    fila: int
    nit: str
    nombre: str
    ok: bool
    mensaje: str


class RespuestaImportacion(BaseModel):
    total_filas: int
    importados: int
    omitidos: int
    errores: int
    detalle: list[FilaResultado]


# ── GET /admin/plantilla-declarantes ──────────────────────────────────────────

@router.get("/plantilla-declarantes")
def descargar_plantilla():
    """Devuelve un .xlsx con las columnas esperadas y una fila de ejemplo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Declarantes"
    ws.append(COLUMNAS_PLANTILLA)
    ws.append([
        "79512345", "4", "García", "Carlos", "empleado",
        450_000_000, 120_000_000, 400_000_000, 100_000_000,
    ])

    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(fill_type="solid", fgColor="C96442")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_declarantes.xlsx"},
    )


# ── POST /admin/importar-declarantes ──────────────────────────────────────────

@router.post("/importar-declarantes", response_model=RespuestaImportacion)
async def importar_declarantes(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(requiere_rol(RolUsuario.ADMIN)),
) -> RespuestaImportacion:
    """
    Importa declarantes desde un archivo Excel (.xlsx).
    Toda la lógica de procesamiento vive en declarante_service.importar_declarantes_desde_excel.
    """
    if not archivo.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="El archivo debe ser .xlsx o .xls",
        )

    contenido = await archivo.read()

    try:
        resultado = svc.importar_declarantes_desde_excel(
            db,
            contenido=contenido,
            nombre_archivo=archivo.filename,
            usuario_id=usuario.id,
        )
    except svc.ArchivoInvalidoError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(e))

    return RespuestaImportacion(
        total_filas=resultado.total_filas,
        importados=resultado.importados,
        omitidos=resultado.omitidos,
        errores=resultado.errores,
        detalle=[
            FilaResultado(
                fila=f.fila,
                nit=f.nit,
                nombre=f.nombre,
                ok=f.ok,
                mensaje=f.mensaje,
            )
            for f in resultado.detalle
        ],
    )